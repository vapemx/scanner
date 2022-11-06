# Aquí los datos del equipo
#
#
#
# El siguiente script realizará automáticamente un análisis de todas las urls
# que hayan sido encontradas por otros procesos de la herramienta.
# 
# El análisis será realizado haciendo uso del módulo vt de python, que se
# comunica con la api de VirusTotal.
#
# 

# Módulos de python

import time
import requests

# Módulos de terceros

import vt
from openpyxl.styles import Font
from openpyxl import Workbook


def scan_link(key_value, links):
    try:
        client = vt.Client(key_value)

        # Para la estructura básica del reporte

        book=Workbook()

        dest_filename = 'reporte_analizador_urls.xlsx'

        sheet=book.active

        sheet.title = ('Reporte')

        sheet.merge_cells('C3:J3')

        sheet['C3'].font = Font(bold=True)

        sheet['C3'] = 'URL'

        sheet.merge_cells('K3:L3')

        sheet['K3'].font = Font(bold=True)

        sheet['K3'] = 'Fecha de análisis'

        sheet.merge_cells('M3:N3')

        sheet['M3'].font = Font(bold=True)

        sheet['M3'] = 'Total de análisis'

        sheet.merge_cells('O3:P3')

        sheet['O3'].font = Font(bold=True)

        sheet['O3'] = 'Riesgo'

        analysis_array= []

        date_array= []

        # Para la obtención de los análisis

        # links = ''  <-- En la versión final, este valor se obtendrá del main

        counter = 0

        i = 0

        for link in links:

            if counter==4:

                print('Límite de links alcanzado. El análisis de las url continuará\
        en un minuto.')

                time.sleep(60)

                counter = 0
        
            response = requests.get(link)

            if response.status_code == 200:
            
                link_id = vt.url_id(links[i])
            
                try:
            
                    analysis = client.get_object('/urls/{}'.format(link_id))

                    stats = analysis.last_analysis_stats
        
                    date = analysis.last_analysis_date

                    analysis_array.append(stats)

                    date_array.append(date)
                    print(date_array)

                except:
                
                    print('Sucedió un error al analizar el enlace: ',links[i])
            
                    links.remove(links[i])

                print(links[i])
                i += 1

                counter += 1


        # Para ingresar los datos obtenidos al reporte

        i = 4
        j = 4
        k = 4


        for i in range (4, len(links)+4):
            
            sheet.merge_cells(f'C{i}:J{i}')
            
            sheet[f'C{i}'] = links[i-4]
            
        for j in range (4, len(date_array)+4):

            sheet.merge_cells(f'K{j}:L{j}')
            
            sheet[f'K{j}'] = date_array[j-4]

        for k in range (4,len(analysis_array)+4):
            
            sheet.merge_cells(f'M{k}:N{k}')
            
            harmless = (analysis_array[k-4])['harmless']

            malicious = (analysis_array[k-4])['malicious']

            tot_ans= harmless + malicious

            sheet[f'M{k}'] = tot_ans

            posit = (analysis_array[k-4])['malicious']

            sheet.merge_cells(f'O{k}:P{k}')

            if (posit<=3):
                
                sheet[f'O{k}'] = 'Baja'

            elif (posit>3 and posit<=10):

                sheet[f'O{k}'] = 'Media'

            elif (posit>10):

                sheet[f'O{k}'] = 'Alta'

        book.save(filename = 'reporte de urls.xlsx')

        print('Análisis de links completado. Podrá encontrar el reporte en excel en el')

        print('directorio actual.')
        return True

    except:
        print("Error al analizar links.")
        return False
